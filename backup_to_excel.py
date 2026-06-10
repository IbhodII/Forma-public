# -*- coding: utf-8 -*-
"""
Экспорт workouts.db в Excel (резервная копия).
Используется из CLI и может запускаться отдельно:
    python backup_to_excel.py
"""

from __future__ import annotations

import json
import logging
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Оформление как в «2 - Тренировки.xlsx»
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="A5A5A5")
_HEADER_FONT = Font(bold=True)
_DATE_FMT = "DD.MM.YYYY"
_DATA_COL_WIDTH = 13.0
_DATE_COL_WIDTH = 11.86
_BLOCK_LABEL = "Упражнения"

_METRIC_HEADERS: tuple[tuple[str, str], ...] = (
    ("avg_hr", "Средний пульс"),
    ("calories_chest", "Ккал, пульсометр"),
    ("calories_watch", "Ккал, браслет"),
)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DB_PATH = SCRIPT_DIR / "workouts.db"
BACKUP_DIR = SCRIPT_DIR / "backups"
LAST_BACKUP_MARKER = BACKUP_DIR / "last_backup.txt"
BACKUP_INTERVAL_DAYS = 7

STRENGTH_RAW_SQL = """
    SELECT date, workout_title, exercise, set_number, weight, reps, notes,
           avg_hr, calories_chest, calories_watch
    FROM strength_workouts
    ORDER BY date DESC, workout_title, exercise, set_number
"""

# Порядок листов силовых (как цикл тренировок в дашборде)
STRENGTH_WORKOUT_ORDER = (
    "Бицепс",
    "Грудь",
    "Спина",
    "Плечи",
    "Ноги",
    "Трицепс",
)

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
DAILY_WEIGHT_SQL = """
    SELECT date, weight_kg, body_fat_percent
    FROM daily_weight
    ORDER BY date DESC
"""
STRENGTH_BY_TITLE_SQL = """
    SELECT date, exercise, set_number, weight, reps,
           avg_hr, calories_chest, calories_watch
    FROM strength_workouts
    WHERE workout_title = ?
    ORDER BY date ASC, exercise, set_number
"""

# Кардио: отдельные листы как в «2 - Тренировки.xlsx»
CARDIO_SHEET_ORDER = ("Бассейн", "Велик", "Бег")
CARDIO_TYPE_BY_SHEET: dict[str, str] = {
    "Бассейн": "бассейн",
    "Велик": "вело",
    "Бег": "бег",
}
CARDIO_BY_TYPE_SQL = """
    SELECT date, distance_km, duration_sec, avg_hr, max_hr,
           calories, calories_chest, calories_watch, swolf
    FROM cardio_workouts
    WHERE type = ?
    ORDER BY date ASC
"""

DAILY_WEIGHT_COL_RU: dict[str, str] = {
    "date": "Дата",
    "weight_kg": "Вес, кг",
    "body_fat_percent": "Жир, %",
}

# Подписи колонок замеров тела (как в дашборде)
BODY_COL_RU: dict[str, str] = {
    "date": "Дата",
    "weight_kg": "Вес, кг",
    "body_fat_percent": "Жир, %",
    "muscle_mass_kg": "Мышцы, кг",
    "chest_inhale_cm": "Грудь вдох, см",
    "chest_exhale_cm": "Грудь выдох, см",
    "chest_avg_cm": "Грудь ср., см",
    "bicep_tense_cm": "Бицепс Н, см",
    "bicep_relaxed_cm": "Бицепс Р, см",
    "bicep_avg_cm": "Бицепс ср., см",
    "bicep_left_cm": "Бицепс лев., см",
    "bicep_right_cm": "Бицепс прав., см",
    "calf_tense_cm": "Икры Н, см",
    "calf_relaxed_cm": "Икры Р, см",
    "calf_avg_cm": "Икры ср., см",
    "calf_left_cm": "Икры лев., см",
    "calf_right_cm": "Икры прав., см",
    "thigh_tense_cm": "Бедро Н, см",
    "thigh_relaxed_cm": "Бедро Р, см",
    "thigh_avg_cm": "Бедро ср., см",
    "thigh_left_cm": "Бедро лев., см",
    "thigh_right_cm": "Бедро прав., см",
    "forearm_tense_cm": "Предпл. Н, см",
    "forearm_relaxed_cm": "Предпл. Р, см",
    "forearm_left_cm": "Предпл. лев., см",
    "forearm_right_cm": "Предпл. прав., см",
    "waist_cm": "Талия, см",
    "hips_cm": "Бёдра, см",
    "ankle_cm": "Лодыжка, см",
    "wrist_cm": "Запястье, см",
    "neck_cm": "Шея, см",
}

# Замеры: 5 пустых колонок между «Мышцы, кг» и обхватами (с «Грудь вдох»)
BODY_GAP_COLS = 5
_BODY_EXPORT_BEFORE_GAP = ("date", "weight_kg", "body_fat_percent", "muscle_mass_kg")
_BODY_EXPORT_AFTER_GAP = tuple(k for k in BODY_COL_RU if k not in _BODY_EXPORT_BEFORE_GAP)


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (name,),
    ).fetchone()
    return row is not None


def _format_date_ddmmyyyy(value: Any) -> str:
    """Дата в виде дд.мм.гггг для Excel."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()[:10]
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%d.%m.%Y")
    except Exception:
        pass
    return text


def _prepare_sheet(
    df: pd.DataFrame,
    labels: dict[str, str],
    *,
    date_columns: tuple[str, ...] = ("date",),
) -> pd.DataFrame:
    """Русские заголовки и даты дд.мм.гггг."""
    if df.empty:
        return df
    out = df.copy()
    for col in date_columns:
        if col in out.columns:
            out[col] = out[col].map(_format_date_ddmmyyyy)
    rename = {c: labels[c] for c in out.columns if c in labels}
    extra = {c: c.replace("_", " ") for c in out.columns if c not in labels}
    out = out.rename(columns={**rename, **extra})
    return out


def _read_table(conn: sqlite3.Connection, sql: str, table: str) -> pd.DataFrame:
    if not _table_exists(conn, table):
        return pd.DataFrame()
    return pd.read_sql_query(sql, conn)


def _num_to_plus_part(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        text = str(value).strip()
        return text if text else None
    if num == int(num):
        return str(int(num))
    return str(num).rstrip("0").rstrip(".")


def _first_non_null(series: pd.Series) -> Any:
    for v in series:
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            return v
    return None


def _style_header_cell(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _single_weight_value(group: pd.DataFrame) -> float | int | None:
    """Один вес на упражнение за день (из первого подхода)."""
    weights = group.sort_values("set_number")["weight"].dropna()
    if weights.empty:
        return None
    first = float(weights.iloc[0])
    return int(first) if first == int(first) else first


def _reps_sum_formula(group: pd.DataFrame) -> str | None:
    """Формула Excel: =6+6+6+6 (сумма подходов)."""
    ordered = group.sort_values("set_number")
    parts = [_num_to_plus_part(v) for v in ordered["reps"]]
    parts = [p for p in parts if p]
    if not parts:
        return None
    return "=" + "+".join(parts)


def _parse_iso_date(date_val: Any) -> datetime | None:
    if date_val is None or (isinstance(date_val, float) and pd.isna(date_val)):
        return None
    text = str(date_val).strip()[:10]
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        parsed = pd.to_datetime(date_val, errors="coerce")
        if pd.notna(parsed):
            return parsed.to_pydatetime()
    return None


def _ordered_exercises(conn: sqlite3.Connection, workout_title: str, raw: pd.DataFrame) -> list[str]:
    """Порядок колонок: шаблон из БД, затем новые упражнения из данных."""
    ordered: list[str] = []
    if _table_exists(conn, "workout_exercise_template"):
        row = conn.execute(
            """
            SELECT MAX(effective_from) FROM workout_exercise_template
            WHERE workout_title = ?
            """,
            (workout_title,),
        ).fetchone()
        eff = row[0] if row else None
        if eff:
            rows = conn.execute(
                """
                SELECT exercise FROM workout_exercise_template
                WHERE workout_title = ? AND effective_from = ?
                ORDER BY sort_order, exercise
                """,
                (workout_title, eff),
            ).fetchall()
            ordered = [r[0] for r in rows if r[0]]
    if not raw.empty:
        by_first_date = (
            raw.groupby("exercise")["date"]
            .min()
            .sort_values()
        )
        for ex in by_first_date.index:
            if ex and ex not in ordered:
                ordered.append(str(ex))
    return ordered


def _write_strength_workout_sheet(
    ws: Worksheet,
    conn: sqlite3.Connection,
    workout_title: str,
) -> int:
    """
    Лист как в «2 - Тренировки.xlsx»: строки — даты, пары колонок Вес/Повторы на упражнение.
    """
    raw = pd.read_sql_query(
        STRENGTH_BY_TITLE_SQL,
        conn,
        params=(workout_title,),
    )
    if raw.empty:
        return 0

    exercises = _ordered_exercises(conn, workout_title, raw)
    if not exercises:
        return 0

    n_ex = len(exercises)
    metric_col_start = 2 + 2 * n_ex

    # --- шапка (строки 1–4) ---
    for row_idx in (1, 2, 3, 4):
        _style_header_cell(ws.cell(row_idx, 1, _BLOCK_LABEL))

    for idx, exercise in enumerate(exercises):
        weight_col = 2 + idx * 2
        reps_col = weight_col + 1
        name_cell = ws.cell(1, weight_col, exercise)
        _style_header_cell(name_cell)
        w_lbl = ws.cell(4, weight_col, "Вес")
        r_lbl = ws.cell(4, reps_col, "Повторы")
        _style_header_cell(w_lbl)
        _style_header_cell(r_lbl)

    for offset, (_, label) in enumerate(_METRIC_HEADERS):
        col = metric_col_start + offset
        _style_header_cell(ws.cell(1, col, label))

    # --- данные по датам (снизу вверх в файле = старые сверху, как в дневнике) ---
    data_rows = 0
    row_idx = 5
    for session_date, day_df in raw.groupby("date", sort=False):
        dt = _parse_iso_date(session_date)
        date_cell = ws.cell(row_idx, 1, dt if dt else session_date)
        if dt:
            date_cell.number_format = _DATE_FMT

        for idx, exercise in enumerate(exercises):
            weight_col = 2 + idx * 2
            reps_col = weight_col + 1
            ex_df = day_df[day_df["exercise"] == exercise]
            if ex_df.empty:
                continue
            w_val = _single_weight_value(ex_df)
            if w_val is not None:
                ws.cell(row_idx, weight_col, w_val)
            reps_formula = _reps_sum_formula(ex_df)
            if reps_formula:
                ws.cell(row_idx, reps_col, reps_formula)

        for offset, (field, _) in enumerate(_METRIC_HEADERS):
            col = metric_col_start + offset
            val = _first_non_null(day_df[field]) if field in day_df.columns else None
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                cell = ws.cell(row_idx, col, int(val) if float(val) == int(float(val)) else val)

        row_idx += 1
        data_rows += 1

    ws.column_dimensions["A"].width = _DATE_COL_WIDTH
    for col in range(2, metric_col_start + len(_METRIC_HEADERS)):
        ws.column_dimensions[get_column_letter(col)].width = _DATA_COL_WIDTH

    ws.freeze_panes = "B5"
    return data_rows


def _is_missing_num(value: Any) -> bool:
    return value is None or (isinstance(value, float) and pd.isna(value))


def _duration_min_sec(duration_sec: Any) -> tuple[int | None, int | None]:
    if _is_missing_num(duration_sec):
        return None, None
    total = int(duration_sec)
    return total // 60, total % 60


def _pool_pace_100m(duration_sec: Any, distance_km: Any) -> tuple[int | None, float | None]:
    """Минуты и секунды темпа на 100 м (как в дневнике)."""
    if _is_missing_num(duration_sec) or _is_missing_num(distance_km) or float(distance_km) <= 0:
        return None, None
    pace_sec = float(duration_sec) / (float(distance_km) * 10.0)
    return int(pace_sec // 60), pace_sec % 60


def _pace_per_km_sec(duration_sec: Any, distance_km: Any) -> float | None:
    if _is_missing_num(duration_sec) or _is_missing_num(distance_km) or float(distance_km) <= 0:
        return None
    return float(duration_sec) / float(distance_km)


def _pace_per_km_min_sec(duration_sec: Any, distance_km: Any) -> tuple[int | None, float | None]:
    pace = _pace_per_km_sec(duration_sec, distance_km)
    if pace is None:
        return None, None
    return int(pace // 60), pace % 60


def _write_cardio_pool_sheet(ws: Worksheet, raw: pd.DataFrame) -> int:
    """Бассейн: строки 1–3 — шапка, данные с 4-й."""
    for row_idx in (1, 2, 3):
        _style_header_cell(ws.cell(row_idx, 1, _BLOCK_LABEL))
    _style_header_cell(ws.cell(1, 2, "Расстояние"))
    _style_header_cell(ws.cell(1, 4, "Продолжительность, мин"))
    _style_header_cell(ws.cell(1, 6, "Продолжительность, сек"))
    _style_header_cell(ws.cell(1, 8, "Средний темп на 100 м"))
    _style_header_cell(ws.cell(1, 10, "swolf"))
    _style_header_cell(ws.cell(1, 11, "Ккал, браслет"))
    _style_header_cell(ws.cell(3, 8, "мин"))
    _style_header_cell(ws.cell(3, 9, "сек"))

    data_rows = 0
    row_idx = 4
    for _, row in raw.iterrows():
        dt = _parse_iso_date(row["date"])
        date_cell = ws.cell(row_idx, 1, dt if dt else row["date"])
        if dt:
            date_cell.number_format = _DATE_FMT
        if pd.notna(row.get("distance_km")):
            ws.cell(row_idx, 2, float(row["distance_km"]))
        minutes, seconds = _duration_min_sec(row.get("duration_sec"))
        if minutes is not None:
            ws.cell(row_idx, 4, minutes)
        if seconds is not None:
            ws.cell(row_idx, 6, seconds)
        p_min, p_sec = _pool_pace_100m(row.get("duration_sec"), row.get("distance_km"))
        if p_min is not None:
            ws.cell(row_idx, 8, p_min)
        if p_sec is not None:
            ws.cell(row_idx, 9, p_sec)
        if pd.notna(row.get("swolf")):
            ws.cell(row_idx, 10, int(row["swolf"]))
        kcal = row.get("calories_watch")
        if pd.notna(kcal):
            ws.cell(row_idx, 11, int(kcal))
        row_idx += 1
        data_rows += 1

    ws.column_dimensions["A"].width = _DATE_COL_WIDTH
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = _DATA_COL_WIDTH
    ws.freeze_panes = "B4"
    return data_rows


def _write_cardio_endurance_sheet(ws: Worksheet, raw: pd.DataFrame) -> int:
    """Велик / Бег: пульс макс./ср., темп, ккал пульсометр."""
    for row_idx in (1, 2, 3):
        _style_header_cell(ws.cell(row_idx, 1, _BLOCK_LABEL))
    _style_header_cell(ws.cell(1, 2, "Расстояние"))
    _style_header_cell(ws.cell(1, 4, "Продолжительность, мин"))
    _style_header_cell(ws.cell(1, 6, "Продолжительность, сек"))
    _style_header_cell(ws.cell(1, 8, "Средний темп"))
    _style_header_cell(ws.cell(1, 10, "Пульс"))
    _style_header_cell(ws.cell(1, 12, "Средний темп, сек"))
    _style_header_cell(ws.cell(1, 13, "Ккал, пульсометр"))
    _style_header_cell(ws.cell(3, 8, "мин"))
    _style_header_cell(ws.cell(3, 9, "сек"))
    _style_header_cell(ws.cell(3, 10, "Макс."))
    _style_header_cell(ws.cell(3, 11, "Сред."))

    data_rows = 0
    row_idx = 4
    for _, row in raw.iterrows():
        dt = _parse_iso_date(row["date"])
        date_cell = ws.cell(row_idx, 1, dt if dt else row["date"])
        if dt:
            date_cell.number_format = _DATE_FMT
        if pd.notna(row.get("distance_km")):
            ws.cell(row_idx, 2, float(row["distance_km"]))
        minutes, seconds = _duration_min_sec(row.get("duration_sec"))
        if minutes is not None:
            ws.cell(row_idx, 4, minutes)
        if seconds is not None:
            ws.cell(row_idx, 6, seconds)
        p_min, p_sec = _pace_per_km_min_sec(row.get("duration_sec"), row.get("distance_km"))
        if p_min is not None:
            ws.cell(row_idx, 8, p_min)
        if p_sec is not None:
            ws.cell(row_idx, 9, p_sec)
        pace_sec = _pace_per_km_sec(row.get("duration_sec"), row.get("distance_km"))
        if pace_sec is not None:
            ws.cell(row_idx, 12, pace_sec)
        if pd.notna(row.get("max_hr")):
            ws.cell(row_idx, 10, int(row["max_hr"]))
        if pd.notna(row.get("avg_hr")):
            ws.cell(row_idx, 11, int(row["avg_hr"]))
        kcal = row.get("calories_chest")
        if pd.isna(kcal) or kcal is None:
            kcal = row.get("calories")
        if pd.notna(kcal):
            ws.cell(row_idx, 13, int(kcal))
        row_idx += 1
        data_rows += 1

    ws.column_dimensions["A"].width = _DATE_COL_WIDTH
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = _DATA_COL_WIDTH
    ws.freeze_panes = "B4"
    return data_rows


def write_cardio_sheets(
    book: Workbook,
    conn: sqlite3.Connection,
) -> tuple[int, int, int]:
    """
    Листы Бассейн / Велик / Бег (макет как в дневнике).
    Возвращает (записей в БД, строк-дат на листах, число листов).
    """
    if not _table_exists(conn, "cardio_workouts"):
        return 0, 0, 0

    total_db = conn.execute("SELECT COUNT(*) FROM cardio_workouts").fetchone()[0]
    data_rows = 0
    sheet_count = 0
    for sheet_name in CARDIO_SHEET_ORDER:
        db_type = CARDIO_TYPE_BY_SHEET[sheet_name]
        raw = pd.read_sql_query(CARDIO_BY_TYPE_SQL, conn, params=(db_type,))
        ws = book.create_sheet(sheet_name)
        sheet_count += 1
        if sheet_name == "Бассейн":
            data_rows += _write_cardio_pool_sheet(ws, raw)
        else:
            data_rows += _write_cardio_endurance_sheet(ws, raw)

    return int(total_db), data_rows, sheet_count


def write_body_metrics_sheet(
    book: Workbook,
    conn: sqlite3.Connection,
) -> int:
    """Лист «Замеры тела»: 5 пустых колонок после «Мышцы, кг»."""
    if not _table_exists(conn, "body_metrics"):
        book.create_sheet("Замеры тела")
        return 0

    raw = pd.read_sql_query(
        "SELECT * FROM body_metrics ORDER BY date ASC",
        conn,
    )
    ws = book.create_sheet("Замеры тела")

    col_idx = 1
    for field in _BODY_EXPORT_BEFORE_GAP:
        _style_header_cell(ws.cell(1, col_idx, BODY_COL_RU[field]))
        col_idx += 1
    col_idx += BODY_GAP_COLS
    field_cols: list[tuple[str, int]] = []
    for field in _BODY_EXPORT_AFTER_GAP:
        if field not in raw.columns:
            continue
        _style_header_cell(ws.cell(1, col_idx, BODY_COL_RU[field]))
        field_cols.append((field, col_idx))
        col_idx += 1

    row_idx = 2
    for _, row in raw.iterrows():
        col_idx = 1
        if "date" in raw.columns:
            dt = _parse_iso_date(row["date"])
            date_cell = ws.cell(row_idx, col_idx, dt if dt else row["date"])
            if dt:
                date_cell.number_format = _DATE_FMT
            col_idx += 1
        for field in _BODY_EXPORT_BEFORE_GAP[1:]:
            if field in raw.columns and pd.notna(row.get(field)):
                ws.cell(row_idx, col_idx, float(row[field]))
            col_idx += 1
        col_idx += BODY_GAP_COLS
        for field, excel_col in field_cols:
            val = row.get(field)
            if pd.notna(val):
                ws.cell(row_idx, excel_col, float(val))
        row_idx += 1

    ws.column_dimensions["A"].width = _DATE_COL_WIDTH
    for c in range(2, col_idx):
        ws.column_dimensions[get_column_letter(c)].width = _DATA_COL_WIDTH
    ws.freeze_panes = "B2"
    return len(raw)


def _excel_sheet_name(title: str, used: set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub(" ", str(title).strip())[:31] or "Тренировка"
    base = name
    n = 2
    while name in used:
        suffix = f" {n}"
        name = (base[: 31 - len(suffix)] + suffix).strip()
        n += 1
    used.add(name)
    return name


def write_strength_sheets(
    book: Workbook,
    conn: sqlite3.Connection,
) -> tuple[int, int, int]:
    """
    Силовые листы в книге Excel (макет как в дневнике).
    Возвращает (число подходов в БД, число строк-дат, число листов).
    """
    raw = _read_table(conn, STRENGTH_RAW_SQL, "strength_workouts")
    set_count = len(raw)
    if raw.empty:
        return set_count, 0, 0

    titles_in_db = raw["workout_title"].dropna().unique().tolist()
    ordered_titles = [t for t in STRENGTH_WORKOUT_ORDER if t in titles_in_db]
    ordered_titles += sorted(t for t in titles_in_db if t not in STRENGTH_WORKOUT_ORDER)

    used_names: set[str] = set()
    total_data_rows = 0
    for title in ordered_titles:
        sheet_name = _excel_sheet_name(str(title), used_names)
        ws = book.create_sheet(sheet_name)
        total_data_rows += _write_strength_workout_sheet(ws, conn, str(title))

    return set_count, total_data_rows, len(ordered_titles)


def backup_database(
    db_path: Path | None = None,
    backup_dir: Path | None = None,
) -> dict[str, Any]:
    """
    Экспорт в Excel: силовые как в «2 - Тренировки.xlsx», плюс Кардио, Замеры, Вес.
    """
    db_path = Path(db_path or DEFAULT_DB_PATH)
    backup_dir = Path(backup_dir or BACKUP_DIR)
    result: dict[str, Any] = {
        "ok": False,
        "path": None,
        "at": None,
        "counts": {
            "strength": 0,
            "cardio": 0,
            "body": 0,
            "daily_weight": 0,
        },
        "error": None,
    }

    if not db_path.exists():
        result["error"] = f"База не найдена: {db_path}"
        logger.warning(result["error"])
        return result

    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
    except OSError as err:
        result["error"] = f"Не удалось создать папку резервных копий: {err}"
        logger.warning(result["error"])
        return result

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out_path = backup_dir / f"backup_{stamp}.xlsx"

    try:
        conn = sqlite3.connect(db_path)
        df_daily = _prepare_sheet(
            _read_table(conn, DAILY_WEIGHT_SQL, "daily_weight"),
            DAILY_WEIGHT_COL_RU,
        )

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            book = writer.book
            strength_sets, strength_rows, strength_sheet_count = write_strength_sheets(
                book, conn
            )
            cardio_total, cardio_rows, cardio_sheet_count = write_cardio_sheets(book, conn)
            body_count = write_body_metrics_sheet(book, conn)
            if "Sheet" in book.sheetnames:
                del book["Sheet"]
            df_daily.to_excel(writer, sheet_name="Ежедневный вес", index=False)
        conn.close()

        at = datetime.now().isoformat(timespec="seconds")
        counts = {
            "strength": strength_sets,
            "strength_rows": strength_rows,
            "strength_sheets": strength_sheet_count,
            "cardio": cardio_total,
            "cardio_rows": cardio_rows,
            "cardio_sheets": cardio_sheet_count,
            "body": body_count,
            "daily_weight": int(len(df_daily)),
        }
        write_backup_marker(backup_dir, at, out_path.name, counts)

        result.update(ok=True, path=str(out_path), at=at, counts=counts)
        logger.info("Backup saved: %s (%s)", out_path, counts)
    except Exception as err:
        result["error"] = str(err)
        logger.exception("Backup failed")

    return result


def write_backup_marker(
    backup_dir: Path,
    at_iso: str,
    filename: str,
    counts: dict[str, int],
    marker_path: Path | None = None,
) -> None:
    """Сохраняет дату и статистику последнего успешного бэкапа."""
    marker = marker_path or (backup_dir / "last_backup.txt")
    payload = {
        "at": at_iso,
        "file": filename,
        "counts": counts,
    }
    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
        marker.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError as err:
        logger.warning("Could not write backup marker: %s", err)


def read_last_backup_info(marker_path: Path | None = None) -> dict[str, Any] | None:
    """Читает last_backup.txt (JSON)."""
    marker = marker_path or LAST_BACKUP_MARKER
    if not marker.exists():
        return None
    try:
        raw = marker.read_text(encoding="utf-8").strip()
        if not raw:
            return None
        if raw.startswith("{"):
            return json.loads(raw)
        return {"at": raw.splitlines()[0].strip(), "counts": {}, "file": None}
    except (OSError, json.JSONDecodeError) as err:
        logger.warning("Could not read backup marker: %s", err)
        return None


def backup_is_due(
    interval_days: int = BACKUP_INTERVAL_DAYS,
    marker_path: Path | None = None,
) -> bool:
    """True, если маркера нет или прошло больше interval_days с последнего бэкапа."""
    info = read_last_backup_info(marker_path)
    if not info or not info.get("at"):
        return True
    try:
        last = datetime.fromisoformat(str(info["at"]))
    except ValueError:
        return True
    return datetime.now() - last > timedelta(days=interval_days)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    r = backup_database()
    if r["ok"]:
        print(f"OK: {r['path']} — {r['counts']}")
    else:
        print(f"Error: {r['error']}")
        raise SystemExit(1)
